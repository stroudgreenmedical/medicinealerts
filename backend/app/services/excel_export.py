import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime
from typing import List
import tempfile
from pathlib import Path

from ..models.alert import Alert


class ExcelExportService:
    """Service for exporting alerts to Excel"""
    
    def export_alerts(
        self,
        alerts: List[Alert],
        start_date: datetime,
        end_date: datetime
    ) -> str:
        """
        Export alerts to Excel file
        
        Returns:
            Path to the generated Excel file
        """
        # Create temporary file
        temp_dir = Path(tempfile.gettempdir())
        file_path = temp_dir / f"mhra_alerts_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        # Create workbook
        wb = Workbook()
        
        # Create main sheet
        ws = wb.active
        ws.title = "Alerts"
        
        # Define columns
        columns = [
            "Alert ID", "GOV.UK Reference", "Title", "URL", "Published Date", "Issued Date",
            "Alert Type", "Severity", "Priority", "Message Type", "Medical Specialties",
            "Product Name", "Active Ingredient", "Manufacturer", "Batch Numbers",
            "Expiry Dates", "Therapeutic Area", "Auto Relevance", "Final Relevance",
            "Relevance Reason", "Status", "Assigned To", "Date First Reviewed",
            "Action Required", "EMIS Search Terms", "EMIS Search Completed",
            "EMIS Search Date", "Patients Affected", "Formulary Check", "Stock Check",
            "Stock Action", "Practice Team Notified", "Practice Team Notified Date",
            "Patients Contacted", "Contact Method", "Communication Template",
            "Recalls Completed", "Action Completed Date", "Time to First Review (hrs)",
            "Time to Completion (hrs)", "Evidence Uploaded", "Evidence Links",
            "CQC Reportable", "Notes", "Created At", "Updated At"
        ]
        
        # Add headers
        for col_num, column_title in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_num, value=column_title)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal="center")
        
        # Add data
        for row_num, alert in enumerate(alerts, 2):
            ws.cell(row=row_num, column=1, value=alert.alert_id)
            ws.cell(row=row_num, column=2, value=alert.govuk_reference)
            ws.cell(row=row_num, column=3, value=alert.title)
            ws.cell(row=row_num, column=4, value=alert.url)
            ws.cell(row=row_num, column=5, value=alert.published_date.strftime("%Y-%m-%d %H:%M") if alert.published_date else "")
            ws.cell(row=row_num, column=6, value=alert.issued_date.strftime("%Y-%m-%d") if alert.issued_date else "")
            ws.cell(row=row_num, column=7, value=alert.alert_type)
            ws.cell(row=row_num, column=8, value=alert.severity.value if alert.severity else "")
            ws.cell(row=row_num, column=9, value=alert.priority.value if alert.priority else "")
            ws.cell(row=row_num, column=10, value=alert.message_type)
            ws.cell(row=row_num, column=11, value=alert.medical_specialties)
            ws.cell(row=row_num, column=12, value=alert.product_name)
            ws.cell(row=row_num, column=13, value=alert.active_ingredient)
            ws.cell(row=row_num, column=14, value=alert.manufacturer)
            ws.cell(row=row_num, column=15, value=alert.batch_numbers)
            ws.cell(row=row_num, column=16, value=alert.expiry_dates)
            ws.cell(row=row_num, column=17, value=alert.therapeutic_area)
            ws.cell(row=row_num, column=18, value=alert.auto_relevance)
            ws.cell(row=row_num, column=19, value=alert.final_relevance)
            ws.cell(row=row_num, column=20, value=alert.relevance_reason)
            ws.cell(row=row_num, column=21, value=alert.status.value if alert.status else "")
            ws.cell(row=row_num, column=22, value=alert.assigned_to)
            ws.cell(row=row_num, column=23, value=alert.date_first_reviewed.strftime("%Y-%m-%d %H:%M") if alert.date_first_reviewed else "")
            ws.cell(row=row_num, column=24, value=alert.action_required)
            ws.cell(row=row_num, column=25, value=alert.emis_search_terms)
            ws.cell(row=row_num, column=26, value="Yes" if alert.emis_search_completed else "No")
            ws.cell(row=row_num, column=27, value=alert.emis_search_date.strftime("%Y-%m-%d") if alert.emis_search_date else "")
            ws.cell(row=row_num, column=28, value=alert.patients_affected_count)
            ws.cell(row=row_num, column=29, value="Yes" if alert.formulary_check else "No" if alert.formulary_check is not None else "")
            ws.cell(row=row_num, column=30, value="Yes" if alert.stock_check else "No" if alert.stock_check is not None else "")
            ws.cell(row=row_num, column=31, value=alert.stock_action)
            ws.cell(row=row_num, column=32, value="Yes" if alert.practice_team_notified else "No")
            ws.cell(row=row_num, column=33, value=alert.practice_team_notified_date.strftime("%Y-%m-%d") if alert.practice_team_notified_date else "")
            ws.cell(row=row_num, column=34, value=alert.patients_contacted)
            ws.cell(row=row_num, column=35, value=alert.contact_method)
            ws.cell(row=row_num, column=36, value=alert.communication_template)
            ws.cell(row=row_num, column=37, value="Yes" if alert.recalls_completed else "No" if alert.recalls_completed is not None else "")
            ws.cell(row=row_num, column=38, value=alert.action_completed_date.strftime("%Y-%m-%d %H:%M") if alert.action_completed_date else "")
            ws.cell(row=row_num, column=39, value=round(alert.time_to_first_review, 2) if alert.time_to_first_review else "")
            ws.cell(row=row_num, column=40, value=round(alert.time_to_completion, 2) if alert.time_to_completion else "")
            ws.cell(row=row_num, column=41, value="Yes" if alert.evidence_uploaded else "No")
            ws.cell(row=row_num, column=42, value=alert.evidence_links)
            ws.cell(row=row_num, column=43, value="Yes" if alert.cqc_reportable else "No")
            ws.cell(row=row_num, column=44, value=alert.notes)
            ws.cell(row=row_num, column=45, value=alert.created_at.strftime("%Y-%m-%d %H:%M") if alert.created_at else "")
            ws.cell(row=row_num, column=46, value=alert.updated_at.strftime("%Y-%m-%d %H:%M") if alert.updated_at else "")
        
        # Add summary sheet
        summary_ws = wb.create_sheet(title="Summary")
        
        # Summary statistics
        total_alerts = len(alerts)
        relevant_alerts = len([a for a in alerts if a.final_relevance == "Relevant"])
        completed_alerts = len([a for a in alerts if a.status and "Completed" in str(a.status.value)])
        
        summary_data = [
            ["Report Period", f"{start_date.strftime('%Y-%m-%d')} to {end_date.strftime('%Y-%m-%d')}"],
            ["Generated", datetime.now().strftime("%Y-%m-%d %H:%M")],
            ["", ""],
            ["Total Alerts", total_alerts],
            ["Relevant Alerts", relevant_alerts],
            ["Completed Alerts", completed_alerts],
            ["Completion Rate", f"{(completed_alerts/relevant_alerts*100):.1f}%" if relevant_alerts > 0 else "N/A"],
        ]
        
        for row_num, row_data in enumerate(summary_data, 1):
            for col_num, value in enumerate(row_data, 1):
                cell = summary_ws.cell(row=row_num, column=col_num, value=value)
                if row_num <= 2 or col_num == 1:
                    cell.font = Font(bold=True)
        
        # Auto-adjust column widths
        for ws in wb.worksheets:
            for column_cells in ws.columns:
                length = max(len(str(cell.value or "")) for cell in column_cells)
                ws.column_dimensions[column_cells[0].column_letter].width = min(length + 2, 50)
        
        # Save workbook
        wb.save(file_path)
        
        return str(file_path)